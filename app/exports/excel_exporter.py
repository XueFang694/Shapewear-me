"""
Exporteur Excel — deux modes d'export disponibles :

  "variants"  (défaut) :
      Une ligne par variante (couleur × taille), identique au comportement
      précédent.  Si un produit n'a pas de variantes, une ligne produit est
      quand même créée.

  "product" :
      Une ligne par produit.  Les couleurs, tailles, SKUs et disponibilités
      sont fusionnées dans une seule cellule chacune, séparées par " | ".
      Le prix affiché est celui du dernier snapshot (prix catalogue global).

Ordre de tri dans les deux modes :
  1. Marque       — alphabétique
  2. Nom          — alphabétique
  3. (mode variants) Couleur → Taille dans l'ordre vêtement standard

Interface publique (inchangée) :
    exporter = ExcelExporter()
    path = exporter.export_from_db(brand_slugs=["spanx"], mode="variants")
    path = exporter.export_from_db(brand_slugs=["spanx"], mode="product")
    path = exporter.export_brand("spanx", mode="product")
    paths = exporter.export_all_brands(mode="variants")
"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Literal

import pandas as pd


ExportMode = Literal["variants", "product"]

# ---------------------------------------------------------------------------
# Ordre de tri des tailles
# ---------------------------------------------------------------------------

_ALPHA_SIZE_ORDER: list[str] = [
    "XXXS", "XXS", "XS", "XS/S",
    "S", "S/M",
    "M", "M/L",
    "L", "L/XL",
    "XL", "XL/XXL",
    "XXL", "2XL", "1X",
    "XXXL", "3XL", "2X",
    "4XL", "3X",
    "5XL", "4X",
    "6XL", "5X",
]

_ALPHA_SIZE_RANK: dict[str, int] = {
    s.upper(): i for i, s in enumerate(_ALPHA_SIZE_ORDER)
}

_NUMERIC_RE = re.compile(r"^(\d+(?:\.\d+)?)([A-Z]*)$", re.IGNORECASE)


def _size_sort_key(size) -> tuple[int, float, str]:
    if pd.isna(size) or size == "":
        return (3, 0.0, "")
    s = str(size).strip().upper()
    if s in _ALPHA_SIZE_RANK:
        return (0, float(_ALPHA_SIZE_RANK[s]), "")
    m = _NUMERIC_RE.match(s)
    if m:
        return (1, float(m.group(1)), m.group(2))
    return (2, 0.0, s)


def _sort_variants_df(df: pd.DataFrame) -> pd.DataFrame:
    sort_cols = [c for c in ("Marque", "Nom", "Couleur") if c in df.columns]
    if not sort_cols:
        return df
    df = df.sort_values(sort_cols, kind="stable", ignore_index=True)
    if "Taille" in df.columns:
        df = df.sort_values(
            sort_cols + ["Taille"],
            key=lambda col: col.map(_size_sort_key) if col.name == "Taille" else col,
            kind="stable",
            ignore_index=True,
        )
    return df


def _sort_product_df(df: pd.DataFrame) -> pd.DataFrame:
    sort_cols = [c for c in ("Marque", "Nom") if c in df.columns]
    if sort_cols:
        df = df.sort_values(sort_cols, kind="stable", ignore_index=True)
    return df


# ---------------------------------------------------------------------------
# Helpers internes
# ---------------------------------------------------------------------------

def _b(v) -> str:
    if v is None:
        return ""
    return "Yes" if v else "No"


def _d(dt) -> str:
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def _comp(json_str: str | None) -> dict:
    if not json_str:
        return {}
    try:
        return json.loads(json_str)
    except (json.JSONDecodeError, TypeError):
        return {}


def _join(values: list, sep: str = " | ") -> str:
    """Joint une liste de valeurs non-vides."""
    return sep.join(str(v) for v in values if v is not None and str(v).strip())


# ---------------------------------------------------------------------------
# Colonnes pour chaque mode
# ---------------------------------------------------------------------------

# Colonnes communes aux deux modes (informations produit)
_PRODUCT_COLS = [
    "Marque", "Nom", "ID Externe", "URL",
    "Catégorie brute", "Famille", "Sous-famille", "Compression", "Zones ciblées",
    "Actif", "Best Seller", "BS depuis",
    "1ère vue", "Dernière vue", "Supprimé le", "Retour stock",
    "Note", "Nb avis",
    "Matière principale", "Doublure",
    "% Nylon", "% Elastane", "% Polyester", "% Cotton",
    "Matière brute",
    "Devise", "Prix", "Prix original", "Remise %", "En promo", "Disponibilité",
]

# Colonnes supplémentaires en mode variants
_VARIANT_COLS = [
    "Couleur", "Couleur canonique", "Taille", "SKU",
    "Dispo variante", "Variante 1ère vue", "Variante der. vue",
    "Variante supp.", "Variante retour",
]

# Colonnes supplémentaires en mode product (valeurs agrégées)
_PRODUCT_AGG_COLS = [
    "Couleurs", "Tailles", "SKUs", "Dispo variantes",
]


# ---------------------------------------------------------------------------
# Exporter principal
# ---------------------------------------------------------------------------

class ExcelExporter:
    """
    Génère un fichier Excel depuis la base de données.

    Deux modes via le paramètre ``mode`` :
        "variants"  → une ligne par variante (comportement historique)
        "product"   → une ligne par produit, attributs fusionnés par " | "

    Interface :
        exporter = ExcelExporter()
        path = exporter.export_from_db(brand_slugs=["spanx"], mode="variants")
        path = exporter.export_brand("spanx", mode="product")
        paths = exporter.export_all_brands(mode="variants")
    """

    def __init__(self, export_dir: Path | None = None) -> None:
        from app.core.config import settings
        self._export_dir = export_dir or settings.EXPORT_DIR
        self._export_dir.mkdir(parents=True, exist_ok=True)

    # ── Interface publique ────────────────────────────────────────────────

    def export_from_db(
        self,
        brand_slugs: list[str] | None = None,
        session_id: int | None = None,
        filename: str | None = None,
        mode: ExportMode = "variants",
    ) -> Path:
        """
        Lit la base de données et génère un fichier Excel.

        Args:
            brand_slugs : filtrer par marques (None = toutes).
            session_id  : utilisé pour le suffixe du nom de fichier.
            filename    : nom de fichier optionnel.
            mode        : "variants" (une ligne/variante) ou "product" (une ligne/produit).

        Returns:
            Chemin du fichier .xlsx créé.
        """
        rows = self._load_rows(brand_slugs, mode=mode)
        return self._write_excel(rows, session_id=session_id, filename=filename, mode=mode)

    def export_brand(
        self,
        brand_slug: str,
        session_id: int | None = None,
        mode: ExportMode = "variants",
    ) -> Path:
        """
        Exporte les données d'une seule marque dans un fichier Excel dédié.

        Args:
            brand_slug : slug de la marque (ex: "spanx").
            session_id : suffixe optionnel du nom de fichier.
            mode       : "variants" ou "product".

        Returns:
            Chemin du fichier .xlsx créé.
        """
        rows = self._load_rows([brand_slug], mode=mode)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        sfx  = f"_session{session_id}" if session_id else ""
        mode_sfx = "_prod" if mode == "product" else "_var"
        filename = f"export_{ts}{sfx}_{brand_slug}{mode_sfx}.xlsx"
        return self._write_excel(rows, session_id=session_id, filename=filename, mode=mode)

    def export_all_brands(
        self,
        session_id: int | None = None,
        mode: ExportMode = "variants",
    ) -> list[Path]:
        """
        Exporte chaque marque dans un fichier Excel distinct.

        Returns:
            Liste des chemins créés, un par marque active.
        """
        from app.storage.database import get_db
        from app.storage.repository import BrandRepository

        with get_db() as db:
            brands = BrandRepository(db).list_active()

        paths: list[Path] = []
        for brand in brands:
            path = self.export_brand(brand.slug, session_id=session_id, mode=mode)
            paths.append(path)
        return paths

    # ── Chargement depuis la DB ───────────────────────────────────────────

    def _load_rows(self, brand_slugs: list[str] | None, mode: ExportMode) -> list[dict]:
        if mode == "product":
            return self._load_rows_product(brand_slugs)
        return self._load_rows_variants(brand_slugs)

    def _load_rows_variants(self, brand_slugs: list[str] | None) -> list[dict]:
        """Mode 'variants' : une ligne par variante couleur × taille."""
        from app.storage.database import get_db
        from app.storage.models import Product
        from app.storage.repository import BrandRepository, SnapshotRepository, VariantRepository

        rows: list[dict] = []

        with get_db() as db:
            brands = BrandRepository(db).list_active()
            if brand_slugs:
                brands = [b for b in brands if b.slug in brand_slugs]

            snap_repo    = SnapshotRepository(db)
            variant_repo = VariantRepository(db)

            for brand in brands:
                products = db.query(Product).filter(Product.brand_id == brand.id).all()

                for product in products:
                    snapshot = snap_repo.get_latest(product.id)
                    variants = variant_repo.get_by_product(product.id)
                    comp     = _comp(product.material_composition_json)

                    prod_price = snapshot.price if snapshot else None
                    prod_orig  = snapshot.original_price if snapshot else None
                    prod_disc  = snapshot.discount_pct if snapshot else None
                    on_sale    = snapshot.on_sale if snapshot else False
                    avail      = snapshot.availability if snapshot else "unknown"
                    currency   = snapshot.currency if snapshot else "USD"

                    try:
                        zones = json.loads(product.target_zones) if product.target_zones else []
                        zones_str = ", ".join(zones)
                    except Exception:
                        zones_str = product.target_zones or ""

                    base = {
                        "Marque":             brand.slug,
                        "Nom":                product.name,
                        "ID Externe":         product.external_id,
                        "URL":                product.url,
                        "Catégorie brute":    product.category_raw or "",
                        "Famille":            product.family or "",
                        "Sous-famille":       product.subfamily or "",
                        "Compression":        product.compression_level or "",
                        "Zones ciblées":      zones_str,
                        "Actif":              _b(product.is_active),
                        "Best Seller":        _b(product.is_best_seller),
                        "BS depuis":          _d(product.best_seller_first_seen),
                        "1ère vue":           _d(product.first_seen),
                        "Dernière vue":       _d(product.last_seen),
                        "Supprimé le":        _d(product.removed_at),
                        "Retour stock":       _d(product.back_in_stock_at),
                        "Note":               product.rating or "",
                        "Nb avis":            product.review_count or "",
                        "Matière principale": product.material_main or "",
                        "Doublure":           product.material_lining or "",
                        "% Nylon":            comp.get("nylon", ""),
                        "% Elastane":         comp.get("elastane", ""),
                        "% Polyester":        comp.get("polyester", ""),
                        "% Cotton":           comp.get("cotton", ""),
                        "Matière brute":      product.material_raw or "",
                        "Devise":             currency,
                        "Prix":               prod_price or "",
                        "Prix original":      prod_orig or "",
                        "Remise %":           prod_disc or "",
                        "En promo":           _b(on_sale),
                        "Disponibilité":      avail,
                    }

                    if variants:
                        for v in variants:
                            v_price = v.price if v.price is not None else prod_price
                            v_orig  = v.original_price if v.original_price is not None else prod_orig
                            v_disc  = ""
                            if v.on_sale and v_price and v_orig:
                                v_disc = round((1 - v_price / v_orig) * 100, 1)
                            elif prod_disc:
                                v_disc = prod_disc

                            row = dict(base)
                            row.update({
                                "Couleur":           v.color or "",
                                "Couleur canonique": v.color_canonical or v.color or "",
                                "Taille":            v.size or "",
                                "SKU":               v.sku or "",
                                "Prix":              v_price if v_price is not None else "",
                                "Prix original":     v_orig if (v.on_sale and v_orig) else "",
                                "Remise %":          v_disc,
                                "En promo":          _b(v.on_sale),
                                "Dispo variante":    _b(v.available),
                                "Variante 1ère vue": _d(v.first_seen),
                                "Variante der. vue": _d(v.last_seen),
                                "Variante supp.":    _d(v.removed_at),
                                "Variante retour":   _d(v.back_in_stock_at),
                            })
                            rows.append(row)
                    else:
                        row = dict(base)
                        row.update({
                            "Couleur":           "",
                            "Couleur canonique": "",
                            "Taille":            "",
                            "SKU":               "",
                            "Dispo variante":    "",
                            "Variante 1ère vue": "",
                            "Variante der. vue": "",
                            "Variante supp.":    "",
                            "Variante retour":   "",
                        })
                        rows.append(row)

        return rows

    def _load_rows_product(self, brand_slugs: list[str] | None) -> list[dict]:
        """Mode 'product' : une ligne par produit, variantes fusionnées."""
        from app.storage.database import get_db
        from app.storage.models import Product
        from app.storage.repository import BrandRepository, SnapshotRepository, VariantRepository

        rows: list[dict] = []

        with get_db() as db:
            brands = BrandRepository(db).list_active()
            if brand_slugs:
                brands = [b for b in brands if b.slug in brand_slugs]

            snap_repo    = SnapshotRepository(db)
            variant_repo = VariantRepository(db)

            for brand in brands:
                products = db.query(Product).filter(Product.brand_id == brand.id).all()

                for product in products:
                    snapshot = snap_repo.get_latest(product.id)
                    variants = variant_repo.get_by_product(product.id)
                    comp     = _comp(product.material_composition_json)

                    prod_price = snapshot.price if snapshot else None
                    prod_orig  = snapshot.original_price if snapshot else None
                    prod_disc  = snapshot.discount_pct if snapshot else None
                    on_sale    = snapshot.on_sale if snapshot else False
                    avail      = snapshot.availability if snapshot else "unknown"
                    currency   = snapshot.currency if snapshot else "USD"

                    try:
                        zones = json.loads(product.target_zones) if product.target_zones else []
                        zones_str = ", ".join(zones)
                    except Exception:
                        zones_str = product.target_zones or ""

                    # ── Agrégation des variantes ──────────────────────────
                    # Trier les variantes pour un affichage cohérent
                    sorted_variants = sorted(
                        variants,
                        key=lambda v: (
                            v.color or "",
                            _size_sort_key(v.size or ""),
                        ),
                    )

                    # Couleurs uniques (ordre d'apparition dans les variantes triées)
                    seen_colors: list[str] = []
                    for v in sorted_variants:
                        color = (v.color_canonical or v.color or "").strip()
                        if color and color not in seen_colors:
                            seen_colors.append(color)

                    # Tailles uniques dans l'ordre standard
                    seen_sizes: list[str] = []
                    for v in sorted(sorted_variants, key=lambda v: _size_sort_key(v.size or "")):
                        size = (v.size or "").strip()
                        if size and size not in seen_sizes:
                            seen_sizes.append(size)

                    # SKUs uniques
                    seen_skus: list[str] = []
                    for v in sorted_variants:
                        sku = (v.sku or "").strip()
                        if sku and sku not in seen_skus:
                            seen_skus.append(sku)

                    # Dispo par variante : "Couleur / Taille : Yes|No"
                    dispo_parts: list[str] = []
                    for v in sorted_variants:
                        label = " / ".join(
                            p for p in [v.color or "", v.size or ""] if p
                        )
                        if label:
                            dispo_parts.append(f"{label}: {_b(v.available)}")

                    row = {
                        "Marque":             brand.slug,
                        "Nom":                product.name,
                        "ID Externe":         product.external_id,
                        "URL":                product.url,
                        "Catégorie brute":    product.category_raw or "",
                        "Famille":            product.family or "",
                        "Sous-famille":       product.subfamily or "",
                        "Compression":        product.compression_level or "",
                        "Zones ciblées":      zones_str,
                        "Actif":              _b(product.is_active),
                        "Best Seller":        _b(product.is_best_seller),
                        "BS depuis":          _d(product.best_seller_first_seen),
                        "1ère vue":           _d(product.first_seen),
                        "Dernière vue":       _d(product.last_seen),
                        "Supprimé le":        _d(product.removed_at),
                        "Retour stock":       _d(product.back_in_stock_at),
                        "Note":               product.rating or "",
                        "Nb avis":            product.review_count or "",
                        "Matière principale": product.material_main or "",
                        "Doublure":           product.material_lining or "",
                        "% Nylon":            comp.get("nylon", ""),
                        "% Elastane":         comp.get("elastane", ""),
                        "% Polyester":        comp.get("polyester", ""),
                        "% Cotton":           comp.get("cotton", ""),
                        "Matière brute":      product.material_raw or "",
                        "Devise":             currency,
                        "Prix":               prod_price or "",
                        "Prix original":      prod_orig or "",
                        "Remise %":           prod_disc or "",
                        "En promo":           _b(on_sale),
                        "Disponibilité":      avail,
                        # Colonnes agrégées
                        "Couleurs":           _join(seen_colors),
                        "Tailles":            _join(seen_sizes),
                        "SKUs":               _join(seen_skus),
                        "Dispo variantes":    _join(dispo_parts, sep="\n"),
                    }
                    rows.append(row)

        return rows

    # ── Écriture du fichier ───────────────────────────────────────────────

    def _write_excel(
        self,
        rows: list[dict],
        session_id: int | None,
        filename: str | None,
        mode: ExportMode = "variants",
    ) -> Path:
        from app.core.logger import get_logger
        log = get_logger(__name__)

        if not filename:
            ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
            sfx     = f"_session{session_id}" if session_id else ""
            mode_sfx = "_prod" if mode == "product" else "_var"
            filename = f"export_{ts}{sfx}{mode_sfx}.xlsx"

        path = self._export_dir / filename

        df = pd.DataFrame(rows)

        # Ordonner les colonnes : d'abord les colonnes communes, puis spécifiques au mode
        if mode == "product":
            ordered_cols = _PRODUCT_COLS + _PRODUCT_AGG_COLS
        else:
            ordered_cols = _PRODUCT_COLS + _VARIANT_COLS

        # Ne garder que les colonnes présentes dans le DataFrame
        existing_cols = [c for c in ordered_cols if c in df.columns]
        # Ajouter d'éventuelles colonnes non listées (sécurité)
        extra_cols = [c for c in df.columns if c not in existing_cols]
        df = df[existing_cols + extra_cols]

        # Trier
        if mode == "product":
            df = _sort_product_df(df)
        else:
            df = _sort_variants_df(df)

        with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
            sheet_name = "Produits" if mode == "product" else "Variantes"
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        _apply_formatting(str(path), sheet_name, mode=mode)

        log.info(
            "Export Excel créé",
            path=str(path),
            rows=len(rows),
            mode=mode,
        )
        return path


# ---------------------------------------------------------------------------
# Formatage openpyxl
# ---------------------------------------------------------------------------

def _apply_formatting(path: str, sheet_name: str, mode: ExportMode = "variants") -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = load_workbook(path)
        for ws in wb.worksheets:
            # En-têtes
            header_fill = PatternFill("solid", start_color="D9D9D9")
            for cell in ws[1]:
                cell.font      = Font(bold=True, name="Arial", size=10)
                cell.fill      = header_fill
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=False
                )

            # Largeurs de colonnes et wrap sur colonnes agrégées (mode product)
            agg_col_names = {"Couleurs", "Tailles", "SKUs", "Dispo variantes"}
            header_map = {
                cell.value: cell.column for cell in ws[1] if cell.value
            }

            for col_idx, col_cells in enumerate(ws.columns, start=1):
                header_val = ws.cell(row=1, column=col_idx).value or ""
                is_agg = header_val in agg_col_names

                if is_agg and mode == "product":
                    # Colonnes agrégées : largeur fixe + wrap activé
                    ws.column_dimensions[get_column_letter(col_idx)].width = 40
                    for cell in col_cells[1:]:  # sauter l'en-tête
                        cell.alignment = Alignment(
                            wrap_text=True, vertical="top"
                        )
                else:
                    max_len = max(
                        (len(str(c.value)) for c in col_cells if c.value is not None),
                        default=8,
                    )
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(
                        max_len + 2, 50
                    )

            # Figer la première ligne
            ws.freeze_panes = "A2"

            # En mode product, auto-hauteur sur les lignes de données
            # (openpyxl ne gère pas l'auto-hauteur ; on fixe une hauteur
            #  raisonnable pour les colonnes à wrap)
            if mode == "product" and any(
                c in header_map for c in agg_col_names
            ):
                for row_idx in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_idx].height = 60

        wb.save(path)
    except Exception:
        pass